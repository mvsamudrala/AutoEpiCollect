import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QStyleFactory, QDesktopWidget, QMainWindow, QApplication, QVBoxLayout, QHBoxLayout, QPushButton, QWidget, QLabel, QComboBox, QPlainTextEdit, QLineEdit, QGridLayout, QCheckBox, QTextEdit, QStackedWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal, QTimer, Qt, QEvent
from PyQt5.QtGui import QFont
import subprocess
import pandas as pd
import numpy as np
import os
from pathlib import Path
import platform
import openpyxl
from io import StringIO
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord
import time
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ChromeOptions
import chromedriver_autoinstaller
# from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import KFold
from sklearn.model_selection import cross_val_score
from functools import partial


# Function to obtain the fasta-formatted gene sequence of interest from UniProt
def get_gene_sequence(target_gene):
    #  Instantiates the web scraping tool and accesses UniProt website
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # Headless means that a new Chrome window doesn't pop up, it accesses Chrome in the background
    # driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(options=options)
    driver.get('https://www.uniprot.org/')

    sleep(1)

    # Anytime By.XPATH is used, it means using the inspect element button to find the html pointer of a certain button
    # or text box or text
    search_box = driver.find_element(By.XPATH,
                                     '/html/body/div[1]/div/div/main/div/div[1]/div/section/form/div[2]/input')
    search_box.send_keys(target_gene)

    sleep(1.5)

    submit_button = driver.find_element(By.XPATH, '//*[@type="submit"]')
    submit_button.click()

    sleep(3.5)

    try:
        choose_table = driver.find_element(By.XPATH, '/html/body/form/div/span/label[2]/input')
    except NoSuchElementException:
        print("No format selector")
    else:
        choose_table.click()

    sleep(0.5)

    view_results_button = driver.find_element(By.XPATH, '/html/body/form/div/section/button')
    view_results_button.click()

    sleep(1)

    filter_human = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/aside/div/ul/li[2]/div/ul/li[1]/a')
    filter_human.click()

    sleep(1)

    gene_button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/main/table/tbody/tr[1]/td[2]/span/a')
    gene_button.click()

    sleep(5)

    download_button = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div/main/div/div[2]/div/button[1]')
    driver.execute_script("arguments[0].scrollIntoView();", download_button)
    driver.execute_script("arguments[0].click();", download_button)

    sleep(0.5)

    fasta_button = driver.find_element(By.XPATH, '/html/body/aside/div[2]/fieldset/label/select/option[2]')
    fasta_button.click()

    download_button2 = driver.find_element(By.XPATH, '/html/body/aside/div[2]/section[1]/a')
    download_button2.click()

    sleep(2)

    child_tab = driver.window_handles[1]
    driver.switch_to.window(child_tab)

    result_fasta = driver.find_element(By.XPATH, '/html/body/pre').text
    with open(f"{target_gene}.fasta", "w") as g:
        g.write(result_fasta)

    for tab in driver.window_handles:
        driver.switch_to.window(tab)
        driver.close()


# Function to create whole mutant fasta genes from each point mutation of interest
def make_mutant_genes(mutant_list, gene_seq, parent_dir):
    for m in mutant_list:
        file_out = parent_dir / "mutant_gene_fastas" / f"{m}.fasta"
        os.makedirs(os.path.dirname(file_out), exist_ok=True)
        with open(file_out, "w") as f:
            for seq_record in SeqIO.parse(open(gene_seq, mode='r'), 'fasta'):
                wild_type = seq_record.seq
                new = m[-1]
                loc = m[:-1]
                loc = loc[1:]
                loc = int(loc) - 1
                new_seq = wild_type[:loc] + new + wild_type[loc + 1:]
                mutant_seq = SeqRecord(seq=new_seq, id=seq_record.id, description=seq_record.description)
                r = SeqIO.write(mutant_seq, f, 'fasta')
                if r != 1:
                    print('Error while writing sequence:  ' + seq_record.id)


# Function that uses the IEDB Tools API to obtain potential epitopes from the mutant gene fasta files, as well as
# the binding affinity of the epitopes to MHC I and II molecules
def get_epitopes_ba(mutant_list, mhc, parent_dir, update):
    if mhc == "I":
        epitopes_dict = {}
        for m in mutant_list:
            target_sequence = ""
            epitope_lengths = ""
            fasta_file = parent_dir / "mutant_gene_fastas" / f"{m}.fasta"
            for seq_record in SeqIO.parse(open(fasta_file, mode='r'), 'fasta'):
                sequence = str(seq_record.seq)
                loc = m[:-1]
                loc = loc[1:]
                loc = int(loc) - 1
                # target_sequence = sequence[loc - 19:loc] + sequence[loc:loc + 20]
                target_sequence = sequence[max(0, loc - 19): min(len(sequence), loc + 20)]
                print(target_sequence)
                with open("MHCI_HLA_input.txt", "r") as h:
                    alleles = h.read() * 2
                    alleles = alleles[:-1]
                epitope_lengths = "9," * 27 + "10," * 26 + "10"

            mhc_i = subprocess.run(["curl", "--data",
                                    f'method=netmhcpan_ba&sequence_text={target_sequence}&allele={alleles}&length={epitope_lengths}',
                                    "http://tools-cluster-interface.iedb.org/tools_api/mhci/"], capture_output=True,
                                   text=True)
            output = mhc_i.stdout
            table = StringIO(output)
            # print(table)
            df = pd.read_table(table, sep=r"\s+")
            df = df.rename(columns={"ic50": "binding affinity (nM)"})
            column_titles = ["allele", "seq_num", "start", "end", "length", "peptide", "core", "icore",
                             "percentile_rank", "binding affinity (nM)"]
            df = df.reindex(columns=column_titles)
            for i in range(df.shape[0]):
                affinity = float(df["binding affinity (nM)"][i])
                if affinity <= 50:
                    df.at[i, "binding score"] = "STRONG"
                elif 50 < affinity <= 500:
                    df.at[i, "binding score"] = "NORMAL"
                elif 500 < affinity <= 5000:
                    df.at[i, "binding score"] = "WEAK"
                else:
                    df.at[i, "binding score"] = "N/A"
            print(df)
            epitopes_dict[m] = df
            print(f"{m} done")
            update.emit(f"{m} done\n")
    else:
        epitopes_dict = {}
        for m in mutant_list:
            target_sequence = ""
            epitope_lengths = ""
            fasta_file = parent_dir / "mutant_gene_fastas" / f"{m}.fasta"
            for seq_record in SeqIO.parse(open(fasta_file, mode='r'), 'fasta'):
                sequence = str(seq_record.seq)
                loc = m[:-1]
                loc = loc[1:]
                loc = int(loc) - 1
                # target_sequence = sequence[loc - 49:loc] + sequence[loc:loc + 50]
                target_sequence = sequence[max(0, loc - 49): min(len(sequence), loc + 50)]
                print(target_sequence)
                with open("MHCII_HLA_input.txt", "r") as h:
                    alleles = h.read()
                    alleles = alleles[:-1]
                epitope_lengths = "15," * 26 + "15"

            mhc_ii = subprocess.run(["curl", "--data",
                                     f'method=netmhciipan&sequence_text={target_sequence}&allele={alleles}&length={epitope_lengths}',
                                     "http://tools-cluster-interface.iedb.org/tools_api/mhcii/"],
                                    capture_output=True, text=True)
            output = mhc_ii.stdout
            table = StringIO(output)
            df = pd.read_table(table, sep=r"\s+")
            df = df.rename(columns={"ic50": "binding affinity (nM)"})
            column_titles = ["allele", "seq_num", "start", "end", "length", "core_peptide", "peptide", "rank",
                             "binding affinity (nM)"]
            df = df.reindex(columns=column_titles)
            for i in range(df.shape[0]):
                affinity = float(df["binding affinity (nM)"][i])
                if affinity <= 50:
                    df.at[i, "binding score"] = "STRONG"
                elif 50 < affinity <= 500:
                    df.at[i, "binding score"] = "NORMAL"
                elif 500 < affinity <= 5000:
                    df.at[i, "binding score"] = "WEAK"
                else:
                    df.at[i, "binding score"] = "N/A"
            print(df)
            epitopes_dict[m] = df
            print(f"{m} done")
            update.emit(f"{m} done\n")
    return epitopes_dict


# Function that only keeps the mutant epitopes from the list of all potential epitopes
def get_mutant_epitopes(mutant_list, mhc, all_epitopes_dict, parent_dir):
    if mhc == "I":
        epitopes_dict = {}
        for m in mutant_list:
            fasta_file = parent_dir / "mutant_gene_fastas" / f"{m}.fasta"
            df = all_epitopes_dict[m].copy()
            for seq_record in SeqIO.parse(open(fasta_file, mode='r'), 'fasta'):
                sequence = seq_record.seq
                loc = m[:-1]
                loc = loc[1:]
                loc = int(loc) - 1
                start_index = max(0, loc - 9)
                end_index = min(len(sequence), loc + 10)
                target_epitope = sequence[start_index:loc] + sequence[loc:end_index]
                print(m + ": " + target_epitope)
                index = 0
                bad_epitopes_indexes = []
                loc += 1
                for pep in df["peptide"]:
                    # print(pep)
                    current_start_index = sequence.find(pep, loc - 10) + 1
                    # print(start_index)
                    current_end_index = current_start_index + len(pep) - 1
                    # print(end_index)
                    if pep not in target_epitope or not current_start_index <= loc <= current_end_index:
                    # if pep not in target_epitope:
                        bad_epitopes_indexes.append(index)
                    else:
                        print(f"{pep}, {current_start_index}, {current_end_index}")
                    index += 1
                # print(bad_epitopes_indexes)
                df_dropped = df.drop(bad_epitopes_indexes).reset_index(drop=True)
                epitopes_dict[m] = df_dropped.reset_index(drop=True)
    else:
        epitopes_dict = {}
        for m in mutant_list:
            fasta_file = parent_dir / "mutant_gene_fastas" / f"{m}.fasta"
            df = all_epitopes_dict[m].copy()
            # print(f"hello: {df['peptide']}")
            for seq_record in SeqIO.parse(open(fasta_file, mode='r'), 'fasta'):
                sequence = seq_record.seq
                loc = m[:-1]
                loc = loc[1:]
                loc = int(loc) - 1
                # print(loc)
                start_index = max(0, loc - 14)
                end_index = min(len(sequence), loc + 15)
                # print(start_index)
                # print(end_index)
                target_epitope = sequence[start_index:loc] + sequence[loc:end_index]
                print(m + ": " + target_epitope)
                bad_epitopes_indexes = []
                loc += 1
                for i in range(df.shape[0]):
                    # start_index = int(df["start"][i])
                    # end_index = int(df["end"][i])
                    pep = df["peptide"][i]
                    # print(pep)
                    # print(loc)
                    current_start_index = sequence.find(pep, loc - 15) + 1
                    # print(current_start_index)
                    current_end_index = current_start_index + len(pep) - 1
                    # print(current_end_index)
                    if pep not in target_epitope or not current_start_index <= loc <= current_end_index:
                    # if pep not in target_epitope:
                        bad_epitopes_indexes.append(i)
                    else:
                        print(f"{pep}, {current_start_index}, {current_end_index}")
                # print(bad_epitopes_indexes)
                df_dropped = df.drop(bad_epitopes_indexes).reset_index(drop=True)
                epitopes_dict[m] = df_dropped
    return epitopes_dict


# Function that makes .txt and .fasta files of peptides created from each point mutation of interest, test
def get_peptides(point_mutants, mut_epitopes_dict, mhc, parent_dir):
    for m in point_mutants:
        file_out = f"{parent_dir}/Sequences/{m}peptides_{mhc}.txt"
        os.makedirs(os.path.dirname(file_out), exist_ok=True)
        df = mut_epitopes_dict[m]
        peptide_set = set()
        for pep in df.loc[:, "peptide"]:
            peptide_set.add(pep)
        with open(file_out, "w") as f_out:
            for pep in peptide_set:
                f_out.write(pep + "\n")
        file_out = parent_dir / "Sequences" / f"{m}peptides_{mhc}.fasta"
        os.makedirs(os.path.dirname(file_out), exist_ok=True)
        with open(file_out, "w") as f_out:
            count = 0
            for pep in peptide_set:
                heading = f">Epitope{count}"
                f_out.write(heading + "\n" + pep + "\n")
                count += 1


# Function that obtains the immunogenicity scores of MHC I peptides using an already downloaded library from the IEDB
def get_local_immunogenicity_mhci(immunogenicity_file, peptide_file, current_df):
    completed_run = subprocess.run(["python", immunogenicity_file, peptide_file], capture_output=True, text=True)
    output = completed_run.stdout
    output = output[output.find("peptide"):]
    table = StringIO(output)
    df = pd.read_table(table, sep=",")

    for i in range(df.shape[0]):
        result = df["score"][i]
        e = df["peptide"][i]
        current_df.loc[current_df["peptide"] == e, "immunogenicity"] = float(result)
    return current_df


# Function that obtain the immunogenicity scores of MHC II peptides by webscraping the IEDB website
# This function is a work in progress as this website is finnicky and slow right now
def get_immunogenicity_mhcii(peptide_list, p, current_df):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome()
    driver.get('http://tools.iedb.org/CD4episcore/')

    elem = WebDriverWait(driver, 60).until(
        ec.presence_of_element_located((By.XPATH, '/html/body/div[3]/form/table/tbody/tr[3]/td[2]/textarea')))

    searchbox = driver.find_element(By.XPATH, '/html/body/div[3]/form/table/tbody/tr[3]/td[2]/textarea')
    print(p)
    searchbox.send_keys(p)

    sleep(1)

    threshold_button = driver.find_element(By.XPATH, '/html/body/div[3]/form/table/tbody/tr[9]/td[2]/select/option[10]')
    threshold_button.click()

    sleep(2)

    submit_button = driver.find_elements(By.XPATH, '/html/body/div[3]/form/table/tbody/tr[12]/th/div/input[1]')
    submit_button[0].click()

    elem = WebDriverWait(driver, 120).until(
        ec.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[1]/h2')))

    for x in range(len(peptide_list)):
        result = driver.find_element(By.XPATH, f'/html/body/div[3]/div[1]/div[3]/table/tbody/tr[{x + 1}]/td[6]').text
        e = driver.find_element(By.XPATH, f'/html/body/div[3]/div[1]/div[3]/table/tbody/tr[{x + 1}]/td[3]').text
        current_df.loc[current_df["peptide"] == e, "immunogenicity"] = float(result)

    driver.close()
    return current_df


# Function that obtains the antigenicity scores of both MHC I and II peptides.
# This function is a work in progress as VaxiJen v2.0 has a human verification loop that is blocking automation
def get_antigenicity(peptide_list, peptide_fasta, current_df):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(options=options)
    driver.get('http://www.ddg-pharmfac.net/vaxijen/VaxiJen/VaxiJen.html')

    file_box = driver.find_element(By.XPATH, '//input[@type="FILE"]')
    file_box.send_keys(str(peptide_fasta))

    organism_button = driver.find_element(By.XPATH,
                                          '/html/body/div/table/tbody/tr[4]/td[3]/form/table/tbody/tr[2]/td[2]/p/select/option[3]')
    organism_button.click()

    submit_button = driver.find_elements(By.XPATH,
                                         '/html/body/div/table/tbody/tr[4]/td[3]/form/table/tbody/tr[3]/td[2]/input[1]')
    submit_button[0].click()

    sleep(0.5)

    for x in range(len(peptide_list)):
        result = driver.find_element(By.XPATH,
                                     f'/html/body/div/table/tbody/tr[4]/td[3]/table/tbody/tr/td/b[{3 * (x + 1)}]').text
        e = driver.find_element(By.XPATH,
                                f'/html/body/div/table/tbody/tr[4]/td[3]/table/tbody/tr/td/font[{2 * (x + 1)}]').text
        current_df.loc[current_df["peptide"] == e, "antigenicity"] = float(result)

    driver.close()
    return current_df


# Function that obtains the allergenicity of MHC class I epitopes using web scraping
def get_allergenicity_algpred(peptide_list, pf, current_df):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(options=options)
    driver.get('https://webs.iiitd.edu.in/raghava/algpred2/batch.html')

    text_box = driver.find_element(By.XPATH,
                                   '/html/body/header/div[3]/section/form/table/tbody/tr/td/font/p/font[1]/textarea')
    text_box.send_keys(pf)

    threshold_value = driver.find_element(By.XPATH,
                                          '/html/body/header/div[3]/section/form/table/tbody/tr/td/font/font/p[3]/font/font[1]/b/select/option[10]')
    threshold_value.click()

    sleep(0.5)

    submit_button = driver.find_elements(By.XPATH,
                                         '/html/body/header/div[3]/section/form/table/tbody/tr/td/font/font/p[3]/font/font[2]/input[2]')
    submit_button[0].click()

    elem = WebDriverWait(driver, 300).until(
        ec.presence_of_element_located((By.XPATH, '/html/body/header/div[3]/main/h1/strong/font/b')))

    for x in range(len(peptide_list)):
        result = driver.find_element(By.XPATH,
                                     f'/html/body/header/div[3]/main/div/table[2]/tbody/tr[{x + 1}]/td[5]').text
        current_df.loc[current_df["peptide"] == peptide_list[x], "allergenicity"] = float(result)

    driver.close()
    return current_df


# Function that obtains the allergenicity of MHC Class II epitopes using webscraping
def get_allergenicity_netallergen(peptide_list, pf, current_df, peptide_fasta):
    input_fasta = pf + ">Epitope119" + "\n" + "TIETLMLLALIAAAA"

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome()
    driver.get('https://services.healthtech.dtu.dk/services/NetAllergen-1.0/')

    try:
        accept_cookies = driver.find_element(By.XPATH, '/html/body/div[5]/div[4]/div[2]')
    except NoSuchElementException:
        print("No cookies")
    else:
        accept_cookies.click()

    sleep(3)

    # text_box = driver.find_element(By.XPATH, '/html/body/main/div/div[3]/div/div[2]/div[1]/form/table/tbody/tr[1]/td/textarea')
    # text_box = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/form/table/tbody/tr[1]/td/textarea')
    # driver.execute_script("arguments[0].scrollIntoView();", text_box)
    # text_box.send_keys(input_fasta)

    file_box = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/form/table/tbody/tr[1]/td/input')
    driver.execute_script("arguments[0].scrollIntoView();", file_box)
    file_box.send_keys(str(peptide_fasta))

    sleep(1)

    # blast_button = driver.find_element(By.XPATH, '/html/body/main/div/div[3]/div/div[2]/div[1]/form/table/tbody/tr[4]/td/input')
    # blast_button = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/form/table/tbody/tr[4]/td/input')
    # driver.execute_script("arguments[0].click();", blast_button)

    # submit_button = driver.find_element(By.XPATH, '/html/body/main/div/div[3]/div/div[2]/div[1]/form/table/tbody/tr[5]/td/input[1]')
    submit_button = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/form/table/tbody/tr[5]/td/input[1]')
    driver.execute_script("arguments[0].click();", submit_button)

    elem = WebDriverWait(driver, 10000).until(
        ec.presence_of_element_located((By.XPATH, '/html/body/font/table/tbody/tr/td[3]/h2')))

    output = driver.find_element(By.XPATH, '/html/body/pre').text
    output = output[:output.find("Explain")]
    table = StringIO(output)
    df = pd.read_table(table, sep=r"\s+")

    count = 0
    for pep in peptide_list:
        result = df["Score_60F"][count]
        current_df.loc[current_df["peptide"] == pep, "allergenicity"] = float(result)
        count += 1
    return current_df


# Function that access ProtParam and obtains the half-life, instability, aliphatic index, isoelectric point, and
# GRAVY score of peptides. Only half-life and instability are used for filtering epitopes.
def get_protparam(peptide_list, h, ins, ali, iso, g, current_df):
    for e in peptide_list:
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        # driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
        driver = webdriver.Chrome(options=options)
        driver.get('https://web.expasy.org/protparam/')

        # searchbox = driver.find_element(By.XPATH, '//*[@id="sib_body"]/form/textarea')
        searchbox = driver.find_element(By.XPATH, '/html/body/main/div/form/textarea')
        searchbox.send_keys(e)

        # submitButton = driver.find_elements(By.XPATH, '/html/body/div[2]/div[2]/form/p[1]/input[2]')
        submitButton = driver.find_element(By.XPATH, '/html/body/main/div/form/input[3]')
        submitButton.click()

        # results = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/pre[2]').text
        results = driver.find_element(By.XPATH, '/html/body/main/div/pre[2]').text
        pi = results[
             results.find('Theoretical pI: ') + 16:results.find('\n', results.find('Theoretical pI: ') + 16, -1)]
        half_life = results[results.find('The estimated half-life is: ') + 28:results.find('hours', results.find(
            'The estimated half-life is: ') + 28, -1) - 1]
        instability = results[results.find('The instability index (II) is computed to be ') + 45:results.find('\n',
                                                                                                              results.find(
                                                                                                                  'The instability index (II) is computed to be ') + 45,
                                                                                                              -1)]
        alipathy = results[
                   results.find('Aliphatic index: ') + 17:results.find('\n', results.find('Aliphatic index: ') + 17,
                                                                       -1)]
        gravy = results[results.find('Grand average of hydropathicity (GRAVY):') + 40:]

        if h:
            if ">" not in half_life:
                current_df.loc[current_df["peptide"] == e, "half-life"] = float(half_life)
            else:
                current_df.loc[current_df["peptide"] == e, "half-life"] = float(half_life[half_life.find(">") + 1:])
        if ins:
            current_df.loc[current_df["peptide"] == e, "instability"] = float(instability)
        if iso:
            current_df.loc[current_df["peptide"] == e, "isoelectric point"] = float(pi)
        if ali:
            current_df.loc[current_df["peptide"] == e, "aliphatic index"] = float(alipathy)
        if g:
            current_df.loc[current_df["peptide"] == e, "GRAVY score"] = float(gravy)

        driver.close()
    return current_df


# Obtains the toxicity of peptides using webscraping
def get_toxicity(peptide_list, pf, current_df):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(options=options)
    driver.get('https://webs.iiitd.edu.in/raghava/toxinpred/multi_submit.php')

    submission_box = driver.find_element(By.XPATH, '//*[@id="input_box"]')
    submission_box.send_keys(pf)

    submit_button = driver.find_elements(By.XPATH,
                                         '/html/body/table[2]/tbody/tr/td/form/fieldset/table[2]/tbody/tr[3]/td/input[2]')
    submit_button[0].click()

    elem = WebDriverWait(driver, 60).until(
        ec.presence_of_element_located((By.XPATH, '/html/body/div[2]/table/thead/tr[1]/td[1]')))

    for x in range(len(peptide_list)):
        result = driver.find_element(By.XPATH, f'/html/body/div[2]/table/tbody/tr[{x + 1}]/td[4]').text
        e = driver.find_element(By.XPATH, f'/html/body/div[2]/table/tbody/tr[{x + 1}]/td[2]/a').text
        current_df.loc[current_df["peptide"] == e, "toxicity"] = result

    driver.close()
    return current_df


# Obtains the IFN-gamma release of epitopes. IFN-gamma only used for filtering MHC Class II epitopes.
def get_ifn(peptide_list, pf, current_df):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    # driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(options=options)
    driver.get('https://webs.iiitd.edu.in/raghava/ifnepitope/predict.php')

    submission_box = driver.find_element(By.XPATH, '//*[@name="sequence"]')
    submission_box.send_keys(pf)

    submit_button = driver.find_elements(By.XPATH, '//*[@type="submit"]')
    submit_button[0].click()

    elem = WebDriverWait(driver, 300).until(ec.presence_of_element_located((By.XPATH, '//*[@class="sorting_asc"]')))

    entry_number = driver.find_element(By.XPATH,
                                       '/html/body/div/div/div[3]/div/div/div/div/div[1]/label/select/option[4]')
    entry_number.click()

    for x in range(len(peptide_list)):
        result = driver.find_element(By.XPATH,
                                     f'/html/body/div/div/div[3]/div/div/div/div/table/tbody/tr[{x + 1}]/td[5]').text
        e = driver.find_element(By.XPATH,
                                f'/html/body/div/div/div[3]/div/div/div/div/table/tbody/tr[{x + 1}]/td[3]/a').text
        current_df.loc[current_df["peptide"] == e, "IFNg"] = result

    driver.close()
    return current_df


# Applies a z-score and min-max normalization on immunogenicity, antigenicity, and allergenicity data.
# Binding affinity data has a log transformation applied to it
def normalize_data(df_collected_epitopes, mhc):
    immunogenicity_name = "immunogenicity"
    antigenicity_name = "antigenicity"
    allergenicity_name = "allergenicity"
    immunogenicity_mean = df_collected_epitopes[immunogenicity_name].mean()
    immunogenicity_std = df_collected_epitopes[immunogenicity_name].std()
    antigenicity_mean = df_collected_epitopes[antigenicity_name].mean()
    antigenicity_std = df_collected_epitopes[antigenicity_name].std()
    allergenicity_mean = df_collected_epitopes[allergenicity_name].mean()
    allergenicity_std = df_collected_epitopes[allergenicity_name].std()
    for i in range(df_collected_epitopes.shape[0]):
        immunogenicity_value = df_collected_epitopes[immunogenicity_name][i]
        antigenicity_value = df_collected_epitopes[antigenicity_name][i]
        allergenicity_value = df_collected_epitopes[allergenicity_name][i]
        new_immunogenicity_value = (immunogenicity_value - immunogenicity_mean) / immunogenicity_std
        new_antigenicity_value = (antigenicity_value - antigenicity_mean) / antigenicity_std
        new_allergenicity_value = (allergenicity_value - allergenicity_mean) / allergenicity_std
        df_collected_epitopes.at[i, "norm immunogenicity"] = new_immunogenicity_value
        df_collected_epitopes.at[i, "norm antigenicity"] = new_antigenicity_value
        df_collected_epitopes.at[i, "norm allergenicity"] = new_allergenicity_value

    immunogenicity_name = "norm immunogenicity"
    antigenicity_name = "norm antigenicity"
    allergenicity_name = "norm allergenicity"
    immunogenicity_max = df_collected_epitopes[immunogenicity_name].max()
    immunogenicity_min = df_collected_epitopes[immunogenicity_name].min()
    antigenicity_max = df_collected_epitopes[antigenicity_name].max()
    antigenicity_min = df_collected_epitopes[antigenicity_name].min()
    allergenicity_max = df_collected_epitopes[allergenicity_name].max()
    allergenicity_min = df_collected_epitopes[allergenicity_name].min()
    for i in range(df_collected_epitopes.shape[0]):
        immunogenicity_value = df_collected_epitopes[immunogenicity_name][i]
        antigenicity_value = df_collected_epitopes[antigenicity_name][i]
        allergenicity_value = df_collected_epitopes[allergenicity_name][i]
        if mhc == "I":
            new_immunogenicity_value = (immunogenicity_value - immunogenicity_min) / (
                        immunogenicity_max - immunogenicity_min)
        else:
            new_immunogenicity_value = (immunogenicity_value - immunogenicity_max) / (
                        immunogenicity_min - immunogenicity_max)
        new_antigenicity_value = (antigenicity_value - antigenicity_min) / (antigenicity_max - antigenicity_min)
        new_allergenicity_value = (allergenicity_value - allergenicity_max) / (allergenicity_min - allergenicity_max)
        df_collected_epitopes.at[i, immunogenicity_name] = new_immunogenicity_value
        df_collected_epitopes.at[i, antigenicity_name] = new_antigenicity_value
        df_collected_epitopes.at[i, allergenicity_name] = new_allergenicity_value
    return df_collected_epitopes


# Trains the logistic regression model and uses it for the scoring function. Scores the epitopes and ranks them.
def apply_scoring_function(df_normalized_epitopes, mhc):
    if mhc == "I":
        training_df = pd.read_csv("refactored_trainingset_cd8.csv")
        training_df["logBindingAffinity"] = np.log(training_df["Binding Affinity"])
        X = training_df[["norm immunogenicity", "norm antigenicity", "norm allergenicity", "logBindingAffinity"]].values
        y = training_df[["Result"]]
    else:
        training_df = pd.read_csv("refactored_trainingset_cd4.csv")
        training_df["logBindingAffinity"] = np.log(training_df["Binding Affinity"])
        X = training_df[["norm immunogenicity", "norm antigenicity", "norm allergenicity", "logBindingAffinity"]].values
        y = training_df[["Result"]]
    logr = LogisticRegression()
    logr.fit(X, y)

    for i in range(df_normalized_epitopes.shape[0]):
        df_normalized_epitopes["log binding affinity"] = np.log(df_normalized_epitopes["binding affinity (nM)"])
        df_x = df_normalized_epitopes[
            ["norm immunogenicity", "norm antigenicity", "norm allergenicity", "log binding affinity"]].values
        predicted_potential = logr.predict_proba(df_x)[:, 1]
        df_normalized_epitopes["potential"] = predicted_potential
    df_normalized_epitopes_ranked = df_normalized_epitopes.sort_values(by=["potential"], ascending=False)
    return df_normalized_epitopes_ranked


# Takes the top 20 epitopes after ranking as been completed and applies the manual filtration
# using the exclusion criteria.
def get_filtered_epitopes(df_ranked_epitopes, mhc, scoring, h, ins, t, ifn):
    if scoring:
        top_20_df = df_ranked_epitopes.head(20)
    else:
        top_20_df = df_ranked_epitopes
    if mhc == "I":
        if h and ins and t:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40) & (top_20_df["toxicity"] == "Non-Toxin")]
        elif h and ins and not t:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40)]
        elif h and not ins and t:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1) & (top_20_df["toxicity"] == "Non-Toxin")]
        elif h and not ins and not t:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1)]
        elif not h and ins and t:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["instability"] < 40) & (top_20_df["toxicity"] == "Non-Toxin")]
        elif not h and ins and not t:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["instability"] < 40)]
        elif not h and not ins and t:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["toxicity"] == "Non-Toxin")]
        else:
            df_filtered_epitopes = top_20_df.copy()
    else:
        if h and ins and t and ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40) & (
                        top_20_df["toxicity"] == "Non-Toxin") & (top_20_df["IFNg"] == "POSITIVE")]
        elif h and ins and t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40) & (
                    top_20_df["toxicity"] == "Non-Toxin")]
        elif h and ins and not t and ifn:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40) & (top_20_df["IFNg"] == "POSITIVE")]
        elif h and ins and not t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1) & (top_20_df["instability"] < 40)]
        elif h and not ins and t and ifn:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["half-life"] > 1) & (top_20_df["toxicity"] == "Non-Toxin") & (
                            top_20_df["IFNg"] == "POSITIVE")]
        elif h and not ins and t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1) & (top_20_df["toxicity"] == "Non-Toxin")]
        elif h and not ins and not t and ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1) & (top_20_df["IFNg"] == "POSITIVE")]
        elif h and not ins and not t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["half-life"] > 1)]
        elif not h and ins and t and ifn:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["instability"] < 40) & (top_20_df["toxicity"] == "Non-Toxin") & (
                            top_20_df["IFNg"] == "POSITIVE")]
        elif not h and ins and t and not ifn:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["instability"] < 40) & (top_20_df["toxicity"] == "Non-Toxin")]
        elif not h and ins and not t and ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["instability"] < 40) & (top_20_df["IFNg"] == "POSITIVE")]
        elif not h and ins and not t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["instability"] < 40)]
        elif not h and not ins and t and ifn:
            df_filtered_epitopes = top_20_df.loc[
                (top_20_df["toxicity"] == "Non-Toxin") & (top_20_df["IFNg"] == "POSITIVE")]
        elif not h and not ins and t and not ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["toxicity"] == "Non-Toxin")]
        elif not h and not ins and not t and ifn:
            df_filtered_epitopes = top_20_df.loc[(top_20_df["IFNg"] == "POSITIVE")]
        else:
            df_filtered_epitopes = top_20_df.copy()
    return df_filtered_epitopes.reset_index(drop=True)


# Uses PCOptim and PCOptim-CD to obtain the optimized list of filtered epitopes for population coverage analysis.
def get_optimized_epitopes(df_filtered_epitopes, mhc):
    filtered_epitopes_str = ""
    op_df = pd.DataFrame(columns=df_filtered_epitopes.columns)
    if mhc == "I":
        java_file = "PopCoverageOptimization.java"
        for i in range(df_filtered_epitopes.shape[0]):
            allele = df_filtered_epitopes["allele"][i]
            peptide = df_filtered_epitopes["peptide"][i]
            filtered_epitopes_str = f"{filtered_epitopes_str}{allele}\t{peptide}\n"
        filtered_epitopes_str = filtered_epitopes_str.rstrip("\n")
        mhc_i = subprocess.run(["java", java_file, filtered_epitopes_str], capture_output=True, text=True)
        output = mhc_i.stdout.rstrip("\n")
        print(output)
        if output != "":
            peptide_epitopes_list = output.split("\n")
            for combination in peptide_epitopes_list:
                allele_list = combination[combination.find("HLA"):].split(",")
                pep = combination[:combination.find("HLA") - 1]
                for allele in allele_list:
                    index = df_filtered_epitopes.index[
                        (df_filtered_epitopes["allele"] == allele) & (df_filtered_epitopes["peptide"] == pep)][0]
                    op_df.loc[len(op_df)] = df_filtered_epitopes.iloc[index]
    else:
        java_file = "CD4PopCoverageOptimization.java"
        allowed_alleles = ["HLA-DMA", "HLA-DMB", "HLA-DPA1", "HLA-DPB1", "HLA-DQA1", "HLA-DQB1", "HLA-DRB1"]
        for i in range(df_filtered_epitopes.shape[0]):
            allele = df_filtered_epitopes["allele"][i]
            allele_check = allele[:allele.find("*")]
            if allele_check not in allowed_alleles or "/" in allele:
                pass
            else:
                peptide = df_filtered_epitopes["peptide"][i]
                filtered_epitopes_str = f"{filtered_epitopes_str}{allele}\t{peptide}\n"
        filtered_epitopes_str = filtered_epitopes_str.rstrip("\n")
        mhc_ii = subprocess.run(["java", java_file, filtered_epitopes_str], capture_output=True, text=True)
        output = mhc_ii.stdout.rstrip("\n")
        print(output)
        if output != "":
            peptide_epitopes_list = output.split("\n")
            for combination in peptide_epitopes_list:
                allele_list = combination[combination.find("HLA"):].split(",")
                pep = combination[:combination.find("HLA") - 1]
                for allele in allele_list:
                    index = df_filtered_epitopes.index[
                        (df_filtered_epitopes["allele"] == allele) & (df_filtered_epitopes["peptide"] == pep)][0]
                    op_df.loc[len(op_df)] = df_filtered_epitopes.iloc[index]
    return op_df


# Helper function to convert the peptide allele list for input into PCOptim and PCOptim-CD
def convert_to_text(df):
    peptide_allele_dict = {}
    for i in range(df.shape[0]):
        peptide = df["peptide"][i]
        allele = df["allele"][i]
        if peptide not in peptide_allele_dict.keys():
            peptide_allele_dict[peptide] = ""
            peptide_allele_dict[peptide] = f"{peptide_allele_dict[peptide]}{allele},"
        else:
            peptide_allele_dict[peptide] = f"{peptide_allele_dict[peptide]}{allele},"
    file_lines = []
    for peptide in peptide_allele_dict.keys():
        alleles = peptide_allele_dict[peptide]
        alleles = alleles[:-1]
        file_lines.append(f"{peptide}\t{alleles}")
    return file_lines


# Helper function to run the population coverage analysis tool and generate population coverage graphs
def population_coverage_helper(pop_filename, reg, mhc, pep_filename, plot_folder, path):
    completed_run = subprocess.run(
        ["python", pop_filename, "-p", reg, "-c", mhc, "-f", pep_filename, "--plot", plot_folder], capture_output=True,
        text=True)
    output = completed_run.stdout
    print(output)
    output = output[output.find("World"):]
    output_list = output.split("\n")
    output_list = output_list[:19]
    if path == "pop_cov_path":
        output_list.insert(0, "Region\tCoverage\tAverage Hit\tpc90")
    else:
        output_list.insert(0, "Region\tOptimized Coverage\tAverage Hit\tpc90")
    output = "\n".join(output_list)
    table = StringIO(output)
    df = pd.read_table(table, sep="\t")
    return df


# Function to create and organize files for population coverage analysis. Uses function above to run the
# population covarage analysis tool that is downloaded.
def get_population_coverage(df_filtered_epitopes, df_optimized_epitopes, mhc, parent_dir):
    if mhc == "I":
        peptide_allele_filename = f"filtered_epitopes_mhci.txt"
        optimized_peptide_allele_filename = f"optimized_epitopes_mhci.txt"
        filtered_epitopes_txt = convert_to_text(df_filtered_epitopes)
        with open(peptide_allele_filename, "w") as fo:
            for line in filtered_epitopes_txt:
                fo.write(f"{line}\n")
        optimized_epitopes_txt = convert_to_text(df_optimized_epitopes)
        with open(optimized_peptide_allele_filename, "w") as fo:
            for line in optimized_epitopes_txt:
                fo.write(f"{line}\n")
    else:
        peptide_allele_filename = f"filtered_epitopes_mhcii.txt"
        optimized_peptide_allele_filename = f"optimized_epitopes_mhcii.txt"
        filtered_epitopes_txt = convert_to_text(df_filtered_epitopes)
        with open(peptide_allele_filename, "w") as fo:
            for line in filtered_epitopes_txt:
                fo.write(f"{line}\n")
        optimized_epitopes_txt = convert_to_text(df_optimized_epitopes)
        with open(optimized_peptide_allele_filename, "w") as fo:
            for line in optimized_epitopes_txt:
                fo.write(f"{line}\n")
    new_dir = Path(f"{parent_dir}/Population_Coverage_Plots/")
    os.makedirs(new_dir, exist_ok=True)
    pop_coverage_filename = parent_dir / "population_coverage" / "calculate_population_coverage.py"
    plot_output_folder = new_dir / f"Regular_Plots"
    os.makedirs(plot_output_folder, exist_ok=True)
    optimized_plot_output_folder = new_dir / f"Optimized_Plots"
    os.makedirs(optimized_plot_output_folder, exist_ok=True)
    regions = ["World", "East Asia", "Northeast Asia", "South Asia", "Southeast Asia", "Southwest Asia", "Europe",
               "East Africa", "West Africa", "Central Africa", "North Africa", "South Africa", "West Indies",
               "North America", "Central America", "South America", "Oceania"]
    region_string = ",".join(regions)
    regular_results = population_coverage_helper(pop_coverage_filename, region_string, mhc, peptide_allele_filename,
                                                 plot_output_folder, "pop_cov_path")
    optimized_results = population_coverage_helper(pop_coverage_filename, region_string, mhc,
                                                   optimized_peptide_allele_filename, optimized_plot_output_folder,
                                                   "op_pop_cov_path")
    return regular_results, optimized_results


class Worker(QThread):
    update_signal = pyqtSignal(str)
    finished = pyqtSignal()
    run_times = 0
    check_output = False
    ready_to_start = True
    is_running = False
    spreadsheet_output1 = ""
    spreadsheet_output2 = ""

    def __init__(self):
        super().__init__()

    # This function is used to call all the functions above and organize the output data
    def auto_epi_collect(self, gene, gene_file, existing, existing2, mut, i, an, al, ali, g, iso, h, ins, t, ifn,
                         filtering, scoring,
                         pop_cov, mhc_classes):
        self.is_running = True
        self.ready_to_start = False
        start = time.time()
        chromedriver_autoinstaller.install()
        parent_dir = Path(os.getcwd())
        print(parent_dir)
        self.update_signal.emit(f"{parent_dir}\n")
        gene_target = gene
        if gene_target != "":
            print("Getting gene sequence in fasta format...")
            self.update_signal.emit("Getting gene sequence in fasta format...\n")
            get_gene_sequence(gene_target)
            gene_file = f"{gene_target}.fasta"
            self.update_signal.emit(f"Obtained gene sequence\n")
        else:
            print("Obtained gene sequence")
            self.update_signal.emit(f"Obtained gene sequence\n")
        cancer_mutations_dict = {}
        mut = mut.rstrip("\n")
        cancer_mutations_list = mut.split("\n")
        for group in cancer_mutations_list:
            can = group[:group.find(":")]
            mutations = group[group.find(":") + 2:].split(",")
            cancer_mutations_dict[can] = mutations
        mutations_cancer_dict = {}
        df_dict = {}
        epitopes_by_cancer_dict = {}
        df_ranked_dict = {}
        ranked_epitopes_by_cancer_dict = {}
        filtered_dict = {}
        pop_cov_dict = {}
        optimized_pop_cov_dict = {}
        cancers = [cancer for cancer in cancer_mutations_dict.keys()]
        all_mutations = []
        for cancer in cancers:
            for mutation in cancer_mutations_dict[cancer]:
                if mutation not in all_mutations:
                    all_mutations.append(mutation)
                    mutations_cancer_dict[mutation] = [cancer]
                else:
                    mutations_cancer_dict[mutation].append(cancer)
        existing_dict = {}
        if existing != "":
            existing_mhci_wb = openpyxl.load_workbook(existing)
            for mutation in existing_mhci_wb.sheetnames:
                existing_df = pd.read_excel(existing, sheet_name=mutation)
                existing_dict[mutation] = existing_df
                for cancer in existing_df["Cancers"]:
                    if not pd.isna(cancer):
                        if cancer not in cancer_mutations_dict.keys():
                            cancer_mutations_dict[cancer] = [mutation]
                        elif mutation not in cancer_mutations_dict[cancer]:
                            cancer_mutations_dict[cancer].append(mutation)
                        if mutation not in mutations_cancer_dict.keys():
                            mutations_cancer_dict[mutation] = [cancer]
                        elif cancer not in mutations_cancer_dict[mutation]:
                            mutations_cancer_dict[mutation].append(cancer)
        existing_dict2 = {}
        if existing2 != "":
            existing_mhcii_wb = openpyxl.load_workbook(existing2)
            for mutation in existing_mhcii_wb.sheetnames:
                existing_df2 = pd.read_excel(existing2, sheet_name=mutation)
                existing_dict2[mutation] = existing_df2
                for cancer in existing_df2["Cancers"]:
                    if not pd.isna(cancer):
                        if cancer not in cancer_mutations_dict.keys():
                            cancer_mutations_dict[cancer] = [mutation]
                        elif mutation not in cancer_mutations_dict[cancer]:
                            cancer_mutations_dict[cancer].append(mutation)
                        if mutation not in mutations_cancer_dict.keys():
                            mutations_cancer_dict[mutation] = [cancer]
                        elif cancer not in mutations_cancer_dict[mutation]:
                            mutations_cancer_dict[mutation].append(cancer)
        if mhc_classes == "I":
            mhc_classes = ["I"]
        elif mhc_classes == "II":
            mhc_classes = ["II"]
        else:
            mhc_classes = ["I", "II"]
        for mhc_class in mhc_classes:
            print("Making mutant fasta proteins...")
            self.update_signal.emit(f"Making mutant fasta proteins...\n")
            make_mutant_genes(all_mutations, gene_file, parent_dir)
            print("Done making mutant fasta proteins")
            self.update_signal.emit(f"Done making mutant fasta proteins\n")
            print("Obtaining all epitopes and binding affinities...")
            self.update_signal.emit(f"Obtaining all epitopes and binding affinities...\n")
            mutant_gene_all_epitopes_dict = get_epitopes_ba(all_mutations, mhc_class, parent_dir, self.update_signal)
            print("Done generating possible epitopes and binding affinities")
            self.update_signal.emit(f"Done generating possible epitopes and binding affinities\n")
            print("Filtering out unmutated epitopes...")
            self.update_signal.emit(f"Filtering out unmutated epitopes...\n")
            mutant_gene_mut_epitopes_dict = get_mutant_epitopes(all_mutations, mhc_class, mutant_gene_all_epitopes_dict,
                                                                parent_dir)
            print("Done filtering out unmutated epitopes")
            self.update_signal.emit(f"Done filtering out unmutated epitopes\n")
            print("Generating possible peptide sequences for all mutations...")
            self.update_signal.emit(f"Generating possible peptide sequences for all mutations...\n")
            get_peptides(all_mutations, mutant_gene_mut_epitopes_dict, mhc_class, parent_dir)
            print("Done generating all peptide sequences")
            self.update_signal.emit(f"Done generating all peptide sequences\n")
            immunogenicity_file = parent_dir / "immunogenicity" / "predict_immunogenicity.py"
            if mhc_class == "I":
                if existing == "":
                    final_out = "all_variables_mhci.xlsx"
                    self.spreadsheet_output1 = "all_variables_mhci.xlsx"
                else:
                    final_out = "all_variables_mhci_updated.xlsx"
                    self.spreadsheet_output1 = "all_variables_mhci_updated.xlsx"
            else:
                if existing2 == "":
                    final_out = "all_variables_mhcii.xlsx"
                    self.spreadsheet_output2 = f"all_variables_mhcii.xlsx"
                else:
                    final_out = "all_variables_mhcii_updated.xlsx"
                    self.spreadsheet_output2 = "all_variables_mhcii_updated.xlsx"
            with pd.ExcelWriter(final_out, engine='openpyxl') as w:
                if mhc_class == "I":
                    if existing != "":
                        for pm in existing_dict.keys():
                            existing_df = existing_dict[pm]
                            existing_df.drop("Cancers", axis=1, inplace=True)
                            df_dict[pm] = existing_df
                            existing_df.to_excel(w, header=True, index=False, sheet_name=pm)
                            pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                                  index=False,
                                                                                                  sheet_name=pm,
                                                                                                  startcol=
                                                                                                  existing_df.shape[1],
                                                                                                  startrow=0)
                    for pm in all_mutations:
                        print(f"Getting results for {pm}...")
                        self.update_signal.emit(f"Getting results for {pm}...\n")
                        current_df = mutant_gene_mut_epitopes_dict[pm].copy()
                        peptide_fasta = parent_dir / "Sequences" / f"{pm}peptides_{mhc_class}.fasta"
                        with open(peptide_fasta, "r") as f:
                            pf = f.read()
                        peptide_file = parent_dir / "Sequences" / f"{pm}peptides_{mhc_class}.txt"
                        with open(peptide_file, "r") as fp:
                            peptides = [line.strip() for line in fp]
                        if i:
                            print(f"Obtaining {pm} immunogenicity...")
                            self.update_signal.emit(f"Obtaining {pm} immunogenicity...\n")
                            current_df = get_local_immunogenicity_mhci(immunogenicity_file, peptide_file, current_df)
                            print(f"Done obtaining {pm} immunogenicity scores")
                            self.update_signal.emit(f"Done obtaining {pm} immunogenicity scores\n")
                        if an:
                            print(f"Obtaining {pm} antigenicity...")
                            self.update_signal.emit(f"Obtaining {pm} antigenicity...\n")
                            current_df = get_antigenicity(peptides, peptide_fasta, current_df)
                            print(f"Done obtaining {pm} antigenicity scores")
                            self.update_signal.emit(f"Done obtaining {pm} antigenicity scores\n")
                        if al:
                            print(f"Obtaining {pm} allergenicity...")
                            self.update_signal.emit(f"Obtaining {pm} allergenicity...\n")
                            current_df = get_allergenicity_algpred(peptides, pf, current_df)
                            print(f"Done obtaining {pm} allergenicity predictions")
                            self.update_signal.emit(f"Done obtaining {pm} allergenicity predictions\n")
                        if t:
                            print(f"Obtaining {pm} toxicity...")
                            self.update_signal.emit(f"Obtaining {pm} toxicity...\n")
                            current_df = get_toxicity(peptides, pf, current_df)
                            print(f"Done obtaining {pm} toxicity predictions")
                            self.update_signal.emit(f"Done obtaining {pm} toxicity predictions\n")
                        if h or ins or ali or iso or g:
                            print(
                                f"Obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores...")
                            self.update_signal.emit(
                                f"Obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores...\n")
                            current_df = get_protparam(peptides, h, ins, ali, iso, g, current_df)
                            print(
                                f"Done obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores")
                            self.update_signal.emit(
                                f"Done obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores\n")
                        if ifn:
                            print(f"Obtaining {pm} IFN-gamma...")
                            self.update_signal.emit(f"Obtaining {pm} IFN-gamma...\n")
                            current_df = get_ifn(peptides, pf, current_df)
                            print(f"Done obtaining {pm} IFN-gamma predictions")
                            self.update_signal.emit(f"Done obtaining {pm} IFN-gamma predictions\n")
                        df_dict[pm] = current_df
                        current_df.to_excel(w, header=True, index=False, sheet_name=pm)
                        pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                              index=False,
                                                                                              sheet_name=pm,
                                                                                              startcol=current_df.shape[
                                                                                                  1],
                                                                                              startrow=0)
                        print(current_df)
                        print(f"Done obtaining all MHC I epitopes and clinical variables for {pm} mutation")
                        self.update_signal.emit(
                            f"Done obtaining all MHC I epitopes and clinical variables for {pm} mutation\n")
                else:
                    if existing2 != "":
                        for pm in existing_dict2.keys():
                            existing_df2 = existing_dict2[pm]
                            existing_df2.drop("Cancers", axis=1, inplace=True)
                            df_dict[pm] = existing_df2
                            existing_df2.to_excel(w, header=True, index=False, sheet_name=pm)
                            pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                                  index=False,
                                                                                                  sheet_name=pm,
                                                                                                  startcol=
                                                                                                  existing_df2.shape[1],
                                                                                                  startrow=0)
                    for pm in all_mutations:
                        print(f"Getting results for {pm}...")
                        self.update_signal.emit(f"Getting results for {pm}...\n")
                        current_df = mutant_gene_mut_epitopes_dict[pm].copy()
                        peptide_fasta = parent_dir / "Sequences" / f"{pm}peptides_{mhc_class}.fasta"
                        with open(peptide_fasta, "r") as f:
                            pf = f.read()
                        peptide_file = parent_dir / "Sequences" / f"{pm}peptides_{mhc_class}.txt"
                        with open(peptide_file, "r") as fp:
                            p = fp.read()
                        with open(peptide_file, "r") as fp:
                            peptides = [line.strip() for line in fp]
                        if i:
                            print(f"Obtaining {pm} immunogenicity...")
                            self.update_signal.emit(f"Obtaining {pm} immunogenicity...\n")
                            current_df = get_immunogenicity_mhcii(peptides, p, current_df)
                            print(f"Done obtaining {pm} immunogenicity scores")
                            self.update_signal.emit(f"Done obtaining {pm} immunogenicity scores\n")
                        if an:
                            print(f"Obtaining {pm} antigenicity...")
                            self.update_signal.emit(f"Obtaining {pm} antigenicity...\n")
                            current_df = get_antigenicity(peptides, peptide_fasta, current_df)
                            print(f"Done obtaining {pm} antigenicity scores")
                            self.update_signal.emit(f"Done obtaining {pm} antigenicity scores\n")
                        if al:
                            print(f"Obtaining {pm} allergenicity...")
                            self.update_signal.emit(f"Obtaining {pm} allergenicity...\n")
                            current_df = get_allergenicity_netallergen(peptides, pf, current_df, peptide_fasta)
                            print(f"Done obtaining {pm} allergenicity predictions")
                            self.update_signal.emit(f"Done obtaining {pm} allergenicity predictions\n")
                        if t:
                            print(f"Obtaining {pm} toxicity...")
                            self.update_signal.emit(f"Obtaining {pm} toxicity...\n")
                            current_df = get_toxicity(peptides, pf, current_df)
                            print(f"Done obtaining {pm} toxicity predictions")
                            self.update_signal.emit(f"Done obtaining {pm} toxicity predictions\n")
                        if h or ins or ali or iso or g:
                            print(
                                f"Obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores...")
                            self.update_signal.emit(
                                f"Obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores...\n")
                            current_df = get_protparam(peptides, h, ins, ali, iso, g, current_df)
                            print(
                                f"Done obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores")
                            self.update_signal.emit(
                                f"Done obtaining {pm} half-life, instability, isoelectric point, aliphatic index, and/or GRAVY scores\n")
                        if ifn:
                            print(f"Obtaining {pm} IFN-gamma...")
                            self.update_signal.emit(f"Obtaining {pm} IFN-gamma...\n")
                            current_df = get_ifn(peptides, pf, current_df)
                            print(f"Done obtaining {pm} IFN-gamma predictions")
                            self.update_signal.emit(f"Done obtaining {pm} IFN-gamma predictions\n")
                        df_dict[pm] = current_df
                        current_df.to_excel(w, header=True, index=False, sheet_name=pm)
                        pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                              index=False,
                                                                                              sheet_name=pm,
                                                                                              startcol=current_df.shape[
                                                                                                  1],
                                                                                              startrow=0)
                        print(current_df)
                        print(f"Done obtaining all MHC II epitopes and clinical variables for {pm} mutation")
                        self.update_signal.emit(
                            f"Done obtaining all MHC II epitopes and clinical variables for {pm} mutation\n")

            if mhc_class == "I":
                if existing == "":
                    all_epitopes_out = "all_epitopes_by_cancer_mhci.xlsx"
                else:
                    all_epitopes_out = "all_epitopes_by_cancer_mhci_updated.xlsx"
            else:
                if existing2 == "":
                    all_epitopes_out = "all_epitopes_by_cancer_mhcii.xlsx"
                else:
                    all_epitopes_out = "all_epitopes_by_cancer_mhcii_updated.xlsx"
            with pd.ExcelWriter(all_epitopes_out, engine='openpyxl') as w:
                for cancer in cancers:
                    sum_all_epitopes_df = pd.DataFrame()
                    for pm in cancer_mutations_dict[cancer]:
                        current_df = df_dict[pm]
                        sum_all_epitopes_df = pd.concat([sum_all_epitopes_df, current_df],
                                                        ignore_index=True)
                    epitopes_by_cancer_dict[cancer] = sum_all_epitopes_df
                    sum_all_epitopes_df.to_excel(w, header=True, index=False, sheet_name=cancer)
            print("Done obtaining all epitopes and clinical variables for all mutations")
            self.update_signal.emit(f"Done obtaining all epitopes and clinical variables for all mutations\n")

            if scoring:
                if mhc_class == "I":
                    if existing == "":
                        ranked_final_out = "all_variables_ranked_mhci.xlsx"
                    else:
                        ranked_final_out = "all_variables_ranked_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        ranked_final_out = "all_variables_ranked_mhcii.xlsx"
                    else:
                        ranked_final_out = "all_variables_ranked_mhcii_updated.xlsx"
                with pd.ExcelWriter(ranked_final_out, engine='openpyxl') as w:
                    for pm in mutations_cancer_dict.keys():
                        print(f"Normalizing and ranking epitopes for {pm} mutation...")
                        self.update_signal.emit(f"Normalizing and ranking epitopes for {pm} mutation...\n")
                        current_df = df_dict[pm]
                        normalized_df = normalize_data(current_df, mhc_class)
                        ranked_df = apply_scoring_function(normalized_df, mhc_class)
                        df_ranked_dict[pm] = ranked_df
                        ranked_df.to_excel(w, header=True, index=False, sheet_name=pm)
                        pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                              index=False,
                                                                                              sheet_name=pm,
                                                                                              startcol=ranked_df.shape[
                                                                                                  1],
                                                                                              startrow=0)
                        print(f"Done normalizing and ranking epitopes for {pm} mutation")
                        self.update_signal.emit(f"Done normalizing and ranking epitopes for {pm} mutation\n")
                if mhc_class == "I":
                    if existing == "":
                        ranked_epitopes_out = "ranked_epitopes_by_cancer_mhci.xlsx"
                    else:
                        ranked_epitopes_out = "ranked_epitopes_by_cancer_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        ranked_epitopes_out = "ranked_epitopes_by_cancer_mhcii.xlsx"
                    else:
                        ranked_epitopes_out = "ranked_epitopes_by_cancer_mhcii_updated.xlsx"
                with pd.ExcelWriter(ranked_epitopes_out, engine='openpyxl') as w:
                    for cancer in cancers:
                        sum_ranked_epitopes_df = pd.DataFrame()
                        for pm in cancer_mutations_dict[cancer]:
                            ranked_df = df_ranked_dict[pm]
                            sum_ranked_epitopes_df = pd.concat([sum_ranked_epitopes_df, ranked_df],
                                                               ignore_index=True)
                        ranked_epitopes_by_cancer_dict[cancer] = sum_ranked_epitopes_df
                        sum_ranked_epitopes_df.to_excel(w, header=True, index=False, sheet_name=cancer)
                print("Done normalizing and ranking epitopes for all mutations")
                self.update_signal.emit(f"Done normalizing and ranking epitopes for all mutations\n")

            if filtering:
                if mhc_class == "I":
                    if existing == "":
                        top_epitopes_out = "top_epitopes_mhci.xlsx"
                    else:
                        top_epitopes_out = "top_epitopes_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        top_epitopes_out = "top_epitopes_mhcii.xlsx"
                    else:
                        top_epitopes_out = "top_epitopes_mhcii_updated.xlsx"
                with pd.ExcelWriter(top_epitopes_out, engine='openpyxl') as w:
                    for pm in mutations_cancer_dict.keys():
                        print(f"Filtering epitopes for {pm} mutation...")
                        self.update_signal.emit(f"Filtering epitopes for {pm} mutation...\n")
                        if scoring:
                            ranked_df = df_ranked_dict[pm]
                        else:
                            ranked_df = df_dict[pm]
                        filtered_df = get_filtered_epitopes(ranked_df, mhc_class, scoring, h, ins, t, ifn)
                        filtered_dict[pm] = filtered_df
                        filtered_df.to_excel(w, header=True, index=False, sheet_name=pm)
                        pd.DataFrame(mutations_cancer_dict[pm], columns=["Cancers"]).to_excel(w, header=True,
                                                                                              index=False,
                                                                                              sheet_name=pm,
                                                                                              startcol=
                                                                                              filtered_df.shape[1],
                                                                                              startrow=0)
                        print(f"Done filtering epitopes for {pm} mutation")
                        self.update_signal.emit(f"Done filtering epitopes for {pm} mutation\n")
                if mhc_class == "I":
                    if existing == "":
                        top_filtered_epitopes_out = "top_epitopes_by_cancer_mhci.xlsx"
                    else:
                        top_filtered_epitopes_out = "top_epitopes_by_cancer_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        top_filtered_epitopes_out = "top_epitopes_by_cancer_mhcii.xlsx"
                    else:
                        top_filtered_epitopes_out = "top_epitopes_by_cancer_mhcii_updated.xlsx"
                with pd.ExcelWriter(top_filtered_epitopes_out, engine='openpyxl') as w:
                    for cancer in cancers:
                        sum_filtered_epitopes_df = pd.DataFrame()
                        for pm in cancer_mutations_dict[cancer]:
                            filtered_df = filtered_dict[pm]
                            sum_filtered_epitopes_df = pd.concat([sum_filtered_epitopes_df, filtered_df],
                                                                 ignore_index=True)
                        pop_cov_dict[cancer] = sum_filtered_epitopes_df
                        sum_filtered_epitopes_df.to_excel(w, header=True, index=False, sheet_name=cancer)
                print("Done filtering epitopes for all mutations")
                self.update_signal.emit(f"Done filtering epitopes for all mutations\n")

            elif not filtering and scoring:
                for cancer in cancers:
                    sum_filtered_epitopes_df = pd.DataFrame()
                    for pm in cancer_mutations_dict[cancer]:
                        sum_filtered_epitopes_df = pd.concat([sum_filtered_epitopes_df, df_ranked_dict[pm].head(20)],
                                                             ignore_index=True)
                    pop_cov_dict[cancer] = sum_filtered_epitopes_df

            else:
                for cancer in cancers:
                    sum_filtered_epitopes_df = pd.DataFrame()
                    for pm in cancer_mutations_dict[cancer]:
                        sum_filtered_epitopes_df = pd.concat([sum_filtered_epitopes_df, df_dict[pm]], ignore_index=True)
                    pop_cov_dict[cancer] = sum_filtered_epitopes_df

            if pop_cov:
                print(f"Obtaining MHC Class {mhc_class} population coverage results...")
                self.update_signal.emit(f"Obtaining MHC Class {mhc_class} population coverage results...\n")
                if mhc_class == "I":
                    if existing == "":
                        top_op_epitopes_out = "optimized_epitopes_mhci.xlsx"
                    else:
                        top_op_epitopes_out = "optimized_epitopes_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        top_op_epitopes_out = "optimized_epitopes_mhcii.xlsx"
                    else:
                        top_op_epitopes_out = "optimized_epitopes_mhcii_updated.xlsx"
                print(f"Optimizing epitopes...")
                self.update_signal.emit(f"Optimizing epitopes...\n")
                with pd.ExcelWriter(top_op_epitopes_out, engine='openpyxl') as w:
                    sum_filtered_epitopes_df = pd.DataFrame()
                    for cancer in cancers:
                        sum_filtered_epitopes_df = pd.concat([sum_filtered_epitopes_df, pop_cov_dict[cancer]], axis=0)
                    sum_filtered_epitopes_df.reset_index(drop=True, inplace=True)
                    if sum_filtered_epitopes_df.empty:
                        print(
                            f"No filtered MHC Class {mhc_class} epitopes")
                        optimized_df = pd.DataFrame(columns=["No Optimized Epitopes"])
                        optimized_df.to_excel(w, header=True, index=False)
                    else:
                        optimized_df = get_optimized_epitopes(sum_filtered_epitopes_df, mhc_class)
                        optimized_df.to_excel(w, header=True, index=False)
                print("Done optimizing epitopes")
                self.update_signal.emit(f"Done optimizing epitopes\n")

                if mhc_class == "I":
                    if existing == "":
                        population_coverage_out = "population_coverage_results_mhci.xlsx"
                    else:
                        population_coverage_out = "population_coverage_results_mhci_updated.xlsx"
                else:
                    if existing2 == "":
                        population_coverage_out = "population_coverage_results_mhcii.xlsx"
                    else:
                        population_coverage_out = "population_coverage_results_mhcii_updated.xlsx"
                with pd.ExcelWriter(population_coverage_out, engine='openpyxl') as w:
                    if not sum_filtered_epitopes_df.empty:
                        print(f"Obtaining population coverage...")
                        self.update_signal.emit(f"Obtaining population coverage...\n")
                        regular_population_coverage, optimized_population_coverage = get_population_coverage(
                            sum_filtered_epitopes_df, optimized_df, mhc_class, parent_dir)
                        regular_population_coverage.to_excel(w, header=True, index=False)
                        optimized_population_coverage.to_excel(w, header=True, index=False,
                                                               startcol=regular_population_coverage.shape[1] + 1,
                                                               startrow=0)
                    else:
                        no_result_df = pd.DataFrame(columns=["No Population Coverage Results"])
                        no_result_df.to_excel(w, header=True, index=False)
                print(f"Done obtaining MHC Class {mhc_class} population coverage results")
                self.update_signal.emit(f"Done obtaining MHC Class {mhc_class} population coverage results\n")
        end = time.time()
        print(
            f"AutoEpiCollect complete in {end - start} seconds, please click the button to the right to see your epitopes for each mutation")
        self.update_signal.emit(
            f"AutoEpiCollect complete in {end - start} seconds, please click the button to the right to see your epitopes for each mutation\n")
        print("All other outputted Excel spreadsheets will be in the same directory as this program")
        self.update_signal.emit(
            f"All other outputted Excel spreadsheets will be in the same directory as this program\n\n")
        self.run_times += 1
        self.check_output = True
        self.ready_to_start = True
        self.is_running = False

# Don't need to worry about everything below, this is all for the GUI.
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setWindowTitle('AutoEpiCollect')

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        # Create a QVBoxLayout to hold the title QLabel and the QHBoxLayout
        self.main_layout = QVBoxLayout(self.central_widget)

        self.stackedWidget = QStackedWidget(self.central_widget)
        self.main_layout.addWidget(self.stackedWidget)

        self.options_page = QWidget()
        self.output_page = QWidget()

        self.stackedWidget.addWidget(self.options_page)
        self.stackedWidget.addWidget(self.output_page)

        self.options_layout = QVBoxLayout(self.options_page)
        # Create the title QLabel
        self.title_label = QLabel('AutoEpiCollect', self.options_page)
        font = QtGui.QFont()
        font.setPointSize(40)  # Set the font size to 20 points
        self.title_label.setFont(font)
        self.title_label.setAlignment(QtCore.Qt.AlignCenter)  # Set alignment to center
        self.options_layout.addWidget(self.title_label)

        # Create a QHBoxLayout to hold the two QVBoxLayouts
        self.horizontal_layout = QHBoxLayout()

        # Create the first QVBoxLayout
        self.layout1 = QVBoxLayout()

        self.gene_box = QLineEdit(self.options_page)
        self.gene_box.setMinimumHeight(40)
        self.layout1.addWidget(self.gene_box)

        self.gene_label = QLabel('Input Gene Name', self.options_page)
        font = QtGui.QFont()
        font.setPointSize(20)  # Set the font size to 20 points
        self.gene_label.setFont(font)
        self.gene_label.setAlignment(QtCore.Qt.AlignCenter)  # Set alignment to center
        self.layout1.addWidget(self.gene_label)

        # Add QComboBox with three options to layout1
        self.starting_combo = QComboBox(self.options_page)
        self.starting_combo.addItems(['--Collection Option', 'Collect New Data', 'Update Existing Data'])
        self.layout1.addWidget(self.starting_combo)

        self.existing_data_button = QPushButton('Input Existing Class I Data (.xlsx)', self.options_page)
        self.layout1.addWidget(self.existing_data_button)

        self.existing_data_2_button = QPushButton('Input Existing Class II Data (.xlsx)', self.options_page)
        self.layout1.addWidget(self.existing_data_2_button)

        self.cancer_pm_box = QTextEdit(self.options_page)
        self.cancer_pm_box.setMinimumHeight(60)
        self.cancer_pm_box.setPlaceholderText("Cancer: PM1,PM2,PM3,etc.\nExample input:\nGBM: R38H,G106V,E545K\nCRC: H1047R,Y1021K")
        self.cancer_pm_box.setLineWrapMode(QTextEdit.NoWrap)
        self.layout1.addWidget(self.cancer_pm_box)

        self.cancer_pm_label = QLabel("Input Cancer Types & Point Mutations", self.options_page)
        font = QtGui.QFont()
        font.setPointSize(20)  # Set the font size to 20 points
        self.cancer_pm_label.setFont(font)
        self.cancer_pm_label.setAlignment(QtCore.Qt.AlignCenter)
        self.layout1.addWidget(self.cancer_pm_label)

        self.horizontal_layout.addLayout(self.layout1)

        # Create the second QVBoxLayout
        self.layout2 = QVBoxLayout()

        self.gene_button = QPushButton('Input Gene File', self.options_page)
        self.gene_button.setMinimumHeight(40)
        self.layout2.addWidget(self.gene_button)

        self.mhc_combo = QComboBox(self.options_page)
        self.mhc_combo.addItems(['--Choose MHC Class', 'Class I', 'Class II', 'Class I & II'])
        self.layout2.addWidget(self.mhc_combo)

        self.layout2a = QGridLayout()

        self.immune_char_checkbox = QCheckBox("Immune Characteristics", self.options_page)
        self.layout2a.addWidget(self.immune_char_checkbox, 0, 0)

        self.immune_layout = QVBoxLayout()

        self.immunogenicity_checkbox = QCheckBox("Immunogenicity", self.options_page)
        self.immune_layout.addWidget(self.immunogenicity_checkbox)

        self.antigenicity_checkbox = QCheckBox("Antigenicity", self.options_page)
        self.immune_layout.addWidget(self.antigenicity_checkbox)

        self.allergenicity_checkbox = QCheckBox("Allergenicity", self.options_page)
        self.immune_layout.addWidget(self.allergenicity_checkbox)

        self.layout2a.addLayout(self.immune_layout, 1, 0)

        self.physicochem_checkbox = QCheckBox("Physicochemical Properties", self.options_page)
        self.layout2a.addWidget(self.physicochem_checkbox, 0, 1)

        self.physicochemical_layout = QVBoxLayout()

        self.aliphatic_checkbox = QCheckBox("Aliphatic Index", self.options_page)
        self.physicochemical_layout.addWidget(self.aliphatic_checkbox)

        self.gravy_checkbox = QCheckBox("GRAVY Score", self.options_page)
        self.physicochemical_layout.addWidget(self.gravy_checkbox)

        self.pi_checkbox = QCheckBox("Isoelectric Point", self.options_page)
        self.physicochemical_layout.addWidget(self.pi_checkbox)

        self.layout2a.addLayout(self.physicochemical_layout, 1, 1)

        self.layout2.addLayout(self.layout2a)

        self.layout2b = QGridLayout()

        self.epistability_checkbox = QCheckBox("Epitope Stability", self.options_page)
        self.layout2b.addWidget(self.epistability_checkbox, 0, 0)

        self.epistab_layout = QVBoxLayout()

        self.half_life_checkbox = QCheckBox("Half-Life", self.options_page)
        self.epistab_layout.addWidget(self.half_life_checkbox)

        self.instability_checkbox = QCheckBox("Instability Index", self.options_page)
        self.epistab_layout.addWidget(self.instability_checkbox)

        self.layout2b.addLayout(self.epistab_layout, 1, 0)

        #self.tox_ifn_layout = QHBoxLayout()

        self.toxicity_checkbox = QCheckBox("Toxicity", self.options_page)
        self.layout2b.addWidget(self.toxicity_checkbox, 0, 1)

        self.ifn_checkbox = QCheckBox("IFN-γ release", self.options_page)
        self.layout2b.addWidget(self.ifn_checkbox, 1, 1)

        self.layout2.addLayout(self.layout2b)

        self.processes_layout = QHBoxLayout()

        self.scoring_checkbox = QCheckBox("Four-Variable Scoring Function", self.options_page)
        self.processes_layout.addWidget(self.scoring_checkbox)

        self.filter_checkbox = QCheckBox("Variable Filtration", self.options_page)
        self.processes_layout.addWidget(self.filter_checkbox)

        self.layout2.addLayout(self.processes_layout)

        self.pop_cov_checkbox = QCheckBox("Population Coverage Analysis", self.options_page)
        self.layout2.addWidget(self.pop_cov_checkbox)

        self.horizontal_layout.addLayout(self.layout2)
        self.horizontal_layout.setStretchFactor(self.layout1, 1)
        self.horizontal_layout.setStretchFactor(self.layout2, 1)
        self.options_layout.addLayout(self.horizontal_layout)

        # Create a QHBoxLayout for the submit button
        self.submit_layout = QHBoxLayout()

        self.submit_button = QPushButton('Submit', self.options_page)
        self.submit_button.setMinimumSize(180, 60)
        self.submit_layout.addWidget(self.submit_button)
        self.submit_layout.setAlignment(QtCore.Qt.AlignCenter)  # Align the submit button to the center

        self.options_layout.addLayout(self.submit_layout)

        #self.options_page.setLayout(self.options_layout)

        self.output_layout = QVBoxLayout(self.output_page)

        self.output_textbox = QPlainTextEdit(self.output_page)
        self.output_textbox.setMinimumSize(700, 400)
        self.output_textbox.setReadOnly(True)
        self.output_layout.addWidget(self.output_textbox)

        self.output_button = QPushButton("Click for Output", self.output_page)
        self.output_button.setMinimumSize(120, 40)
        self.output_layout.addWidget(self.output_button)

        self.back_button = QPushButton("Back", self.output_page)
        self.back_button.setMinimumSize(180, 60)
        self.output_layout.addWidget(self.back_button)

        self.submit_button.clicked.connect(self.take_text_input)
        self.gene_button.clicked.connect(lambda: self.open_filebox(1))
        self.existing_data_button.clicked.connect(lambda: self.open_filebox(2))
        self.existing_data_2_button.clicked.connect(lambda: self.open_filebox(3))

        self.immune_char_checkbox.stateChanged.connect(self.check_immune_characteristics)
        self.physicochem_checkbox.stateChanged.connect(self.check_physicochemical_properties)
        self.epistability_checkbox.stateChanged.connect(self.check_epitope_stability)

        self.immunogenicity_checkbox.stateChanged.connect(self.check_immunogenicity)
        self.antigenicity_checkbox.stateChanged.connect(self.check_antigenicity)
        self.allergenicity_checkbox.stateChanged.connect(self.check_allergenicity)

        self.aliphatic_checkbox.stateChanged.connect(self.check_aliphatic)
        self.gravy_checkbox.stateChanged.connect(self.check_gravy)
        self.pi_checkbox.stateChanged.connect(self.check_pi)

        self.half_life_checkbox.stateChanged.connect(self.check_half_life)
        self.instability_checkbox.stateChanged.connect(self.check_instability)

        self.scoring_checkbox.stateChanged.connect(lambda: self.immune_char_checkbox.setChecked(True))

        self.back_button.clicked.connect(self.back_home_auto_epi_collect)
        self.output_button.clicked.connect(lambda: self.display_result(self.mhc_combo.currentText()))

        self.gene_file = ""
        self.existing_data = ""
        self.existing_data_2 = ""
        self.data_filename = ""
        self.data_filename_2 = ""

        self.worker = Worker()
        self.worker_thread = QThread()
        self.worker.moveToThread(self.worker_thread)
        self.worker.update_signal.connect(self.write_to_textbox)

        self.updateLayout()

    def resizeEvent(self, event):
        self.updateLayout()

    def updateLayout(self):
        if self.isFullScreen() or self.windowState() & QtCore.Qt.WindowMaximized:
            self.main_layout.setAlignment(QtCore.Qt.AlignCenter)
        else:
            self.main_layout.setAlignment(QtCore.Qt.AlignCenter)

    def check_immune_characteristics(self):
        if self.immune_char_checkbox.isChecked():
            self.immunogenicity_checkbox.setChecked(True)
            self.antigenicity_checkbox.setChecked(True)
            self.allergenicity_checkbox.setChecked(True)
        elif not self.immune_char_checkbox.isChecked():
            self.immunogenicity_checkbox.setChecked(False)
            self.antigenicity_checkbox.setChecked(False)
            self.allergenicity_checkbox.setChecked(False)

    def check_physicochemical_properties(self):
        if self.physicochem_checkbox.isChecked():
            self.aliphatic_checkbox.setChecked(True)
            self.gravy_checkbox.setChecked(True)
            self.pi_checkbox.setChecked(True)
        elif not self.physicochem_checkbox.isChecked():
            self.aliphatic_checkbox.setChecked(False)
            self.gravy_checkbox.setChecked(False)
            self.pi_checkbox.setChecked(False)

    def check_epitope_stability(self):
        if self.epistability_checkbox.isChecked():
            self.half_life_checkbox.setChecked(True)
            self.instability_checkbox.setChecked(True)
        elif not self.epistability_checkbox.isChecked():
            self.half_life_checkbox.setChecked(False)
            self.instability_checkbox.setChecked(False)

    def check_immunogenicity(self):
        antigenicity_check = self.antigenicity_checkbox.isChecked()
        allergenicity_check = self.allergenicity_checkbox.isChecked()
        if not self.immunogenicity_checkbox.isChecked():
            self.scoring_checkbox.setChecked(False)
            if self.immune_char_checkbox.isChecked():
                self.immune_char_checkbox.setChecked(False)
            self.antigenicity_checkbox.setChecked(antigenicity_check)
            self.allergenicity_checkbox.setChecked(allergenicity_check)
        elif self.immunogenicity_checkbox.isChecked():
            if antigenicity_check and allergenicity_check:
                self.immune_char_checkbox.setChecked(True)

    def check_antigenicity(self):
        immunogenicity_check = self.immunogenicity_checkbox.isChecked()
        allergenicity_check = self.allergenicity_checkbox.isChecked()
        if not self.antigenicity_checkbox.isChecked():
            self.scoring_checkbox.setChecked(False)
            if self.immune_char_checkbox.isChecked():
                self.immune_char_checkbox.setChecked(False)
            self.immunogenicity_checkbox.setChecked(immunogenicity_check)
            self.allergenicity_checkbox.setChecked(allergenicity_check)
        elif self.antigenicity_checkbox.isChecked():
            if immunogenicity_check and allergenicity_check:
                self.immune_char_checkbox.setChecked(True)

    def check_allergenicity(self):
        antigenicity_check = self.antigenicity_checkbox.isChecked()
        immunogenicity_check = self.immunogenicity_checkbox.isChecked()
        if not self.allergenicity_checkbox.isChecked():
            self.scoring_checkbox.setChecked(False)
            if self.immune_char_checkbox.isChecked():
                self.immune_char_checkbox.setChecked(False)
            self.antigenicity_checkbox.setChecked(antigenicity_check)
            self.immunogenicity_checkbox.setChecked(immunogenicity_check)
        elif self.allergenicity_checkbox.isChecked():
            if antigenicity_check and immunogenicity_check:
                self.immune_char_checkbox.setChecked(True)

    def check_aliphatic(self):
        gravy_check = self.gravy_checkbox.isChecked()
        pi_check = self.pi_checkbox.isChecked()
        if not self.aliphatic_checkbox.isChecked():
            if self.physicochem_checkbox.isChecked():
                self.physicochem_checkbox.setChecked(False)
            self.gravy_checkbox.setChecked(gravy_check)
            self.pi_checkbox.setChecked(pi_check)
        elif self.aliphatic_checkbox.isChecked():
            if gravy_check and pi_check:
                self.physicochem_checkbox.setChecked(True)

    def check_gravy(self):
        aliphatic_check = self.aliphatic_checkbox.isChecked()
        pi_check = self.pi_checkbox.isChecked()
        if not self.gravy_checkbox.isChecked():
            if self.physicochem_checkbox.isChecked():
                self.physicochem_checkbox.setChecked(False)
            self.aliphatic_checkbox.setChecked(aliphatic_check)
            self.pi_checkbox.setChecked(pi_check)
        elif self.gravy_checkbox.isChecked():
            if aliphatic_check and pi_check:
                self.physicochem_checkbox.setChecked(True)

    def check_pi(self):
        aliphatic_check = self.aliphatic_checkbox.isChecked()
        gravy_check = self.gravy_checkbox.isChecked()
        if not self.pi_checkbox.isChecked():
            if self.physicochem_checkbox.isChecked():
                self.physicochem_checkbox.setChecked(False)
            self.aliphatic_checkbox.setChecked(aliphatic_check)
            self.gravy_checkbox.setChecked(gravy_check)
        elif self.pi_checkbox.isChecked():
            if aliphatic_check and gravy_check:
                self.physicochem_checkbox.setChecked(True)

    def check_half_life(self):
        instability_check = self.instability_checkbox.isChecked()
        if not self.half_life_checkbox.isChecked():
            if self.epistability_checkbox.isChecked():
                self.epistability_checkbox.setChecked(False)
            self.instability_checkbox.setChecked(instability_check)
        elif self.half_life_checkbox.isChecked():
            if instability_check:
                self.epistability_checkbox.setChecked(True)

    def check_instability(self):
        half_life_check = self.half_life_checkbox.isChecked()
        if not self.instability_checkbox.isChecked():
            if self.epistability_checkbox.isChecked():
                self.epistability_checkbox.setChecked(False)
            self.half_life_checkbox.setChecked(half_life_check)
        elif self.instability_checkbox.isChecked():
            if half_life_check:
                self.epistability_checkbox.setChecked(True)

    def display_result(self, mhc):
        os.name = platform.system()
        if self.worker.check_output:
            if mhc == "Class I":
                print(self.worker.spreadsheet_output1)
                if os.name == "Windows":
                    os.startfile(self.worker.spreadsheet_output1)
                elif os.name == "Darwin":
                    subprocess.run(["open", self.worker.spreadsheet_output1])
                elif os.name == "Linux":
                    subprocess.run(["xdg-open", self.worker.spreadsheet_output1])
            elif mhc == "Class II":
                print(self.worker.spreadsheet_output2)
                if os.name == "Windows":
                    os.startfile(self.worker.spreadsheet_output2)
                elif os.name == "Darwin":
                    subprocess.run(["open", self.worker.spreadsheet_output2])
                elif os.name == "Linux":
                    subprocess.run(["xdg-open", self.worker.spreadsheet_output2])
            else:
                print(self.worker.spreadsheet_output1)
                print(self.worker.spreadsheet_output2)
                if os.name == "Windows":
                    os.startfile(self.worker.spreadsheet_output1)
                    os.startfile(self.worker.spreadsheet_output2)
                elif os.name == "Darwin":
                    subprocess.run(["open", self.worker.spreadsheet_output1])
                    subprocess.run(["open", self.worker.spreadsheet_output2])
                elif os.name == "Linux":
                    subprocess.run(["xdg-open", self.worker.spreadsheet_output1])
                    subprocess.run(["xdg-open", self.worker.spreadsheet_output2])

    def take_text_input(self):
        immunogenicity = self.immunogenicity_checkbox.isChecked()
        antigenicity = self.antigenicity_checkbox.isChecked()
        allergenicity = self.allergenicity_checkbox.isChecked()
        aliphatic = self.aliphatic_checkbox.isChecked()
        gravy = self.gravy_checkbox.isChecked()
        isoelectric = self.pi_checkbox.isChecked()
        half_life = self.half_life_checkbox.isChecked()
        instability = self.instability_checkbox.isChecked()
        toxicity = self.toxicity_checkbox.isChecked()
        ifn = self.ifn_checkbox.isChecked()
        filtering = self.filter_checkbox.isChecked()
        scoring_function = self.scoring_checkbox.isChecked()
        population_coverage = self.pop_cov_checkbox.isChecked()
        gene = self.gene_box.text()
        mutations = self.cancer_pm_box.toPlainText()
        if gene == "" and mutations == "" and self.gene_file == "" and self.mhc_combo.currentText() == "--Choose MHC Class":
            self.show_popup(1)
        elif gene == "" and self.gene_file == "" and mutations == "":
            self.show_popup(7)
        elif mutations == "" and self.mhc_combo.currentText() == "--Choose MHC Class":
            self.show_popup(8)
        elif gene == "" and self.gene_file == "" and self.mhc_combo.currentText() == "--Choose MHC Class":
            self.show_popup(9)
        elif gene == "" and self.gene_file == "":
            self.show_popup(2)
        elif mutations == "":
            self.show_popup(3)
        elif self.mhc_combo.currentText() == "--Choose MHC Class":
            self.show_popup(6)
        else:
            if self.mhc_combo.currentText() == "Class I":
                mhc = "I"
            elif self.mhc_combo.currentText() == "Class II":
                mhc = "II"
            else:
                mhc = "I,II"
            if self.starting_combo.currentText() == "Collect New Data":
                if self.existing_data != "" or self.existing_data_2 != "":
                    self.show_popup(10)
                else:
                    self.stackedWidget.setCurrentWidget(self.output_page)
                    if self.worker.ready_to_start and not self.worker.is_running:
                        self.run_auto_epi_collect(gene, self.gene_file, "", "", mutations, immunogenicity, antigenicity,
                                                  allergenicity, aliphatic, gravy, isoelectric, half_life,
                                                  instability, toxicity, ifn, filtering, scoring_function,
                                                  population_coverage, mhc)
            elif self.starting_combo.currentText() == "Update Existing Data":
                if (self.existing_data == "" or ".xlsx" not in self.data_filename) and (
                        self.existing_data_2 == "" or ".xlsx" not in self.data_filename_2):
                    self.show_popup(5)
                elif self.existing_data != "" and mhc == "II":
                    self.show_popup(11)
                elif self.existing_data_2 != "" and mhc == "I":
                    self.show_popup(11)
                elif (self.existing_data != "" and self.existing_data_2 != "") and (mhc == "I" or mhc == "II"):
                    self.show_popup(11)
                else:
                    self.stackedWidget.setCurrentWidget(self.output_page)
                    if self.worker.ready_to_start and not self.worker.is_running:
                        self.run_auto_epi_collect(gene, self.gene_file, self.existing_data, self.existing_data_2, mutations, immunogenicity, antigenicity,
                                                  allergenicity, aliphatic, gravy, isoelectric, half_life,
                                                  instability, toxicity, ifn, filtering, scoring_function,
                                                  population_coverage, mhc)
            else:
                self.show_popup(4)

    def show_popup(self, option):
        msg = QMessageBox()
        msg.setWindowTitle("Alert!")
        if option == 1:
            msg.setText("Please enter a gene or gene file, mutation(s), and an MHC Class.")
            msg.setIcon(QMessageBox.Critical)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 2:
            msg.setText("Please enter a gene or gene file.")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 3:
            msg.setText("Please enter mutation(s).")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 4:
            msg.setText("Please select a collection option.")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 5:
            msg.setText("Please input your existing data in .xlsx format.")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 6:
            msg.setText("Please choose an MHC Class.")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 7:
            msg.setText("Please enter a gene or gene file and mutation(s).")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 8:
            msg.setText("Please enter mutation(s) and choose an MHC Class.")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 9:
            msg.setText("Please enter a gene or gene file and choose an MHC Class")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 10:
            msg.setText("If you want to collect new data, please remove file from existing data boxes")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)
        if option == 11:
            msg.setText("Please select correct MHC Class option for existing data spreadsheet(s)")
            msg.setIcon(QMessageBox.Warning)
            msg.setStandardButtons(QMessageBox.Ok)

        x = msg.exec_()

    def open_filebox(self, option):
        if option == 1:
            self.gene_file = QFileDialog.getOpenFileName()[0]
            self.gene_filename = self.gene_file[self.gene_file.rfind("/") + 1:]
            if self.gene_filename == "":
                self.gene_button.setText("Input Gene File")
            else:
                self.gene_button.setText(self.gene_filename)
        elif option == 2:
            self.existing_data = QFileDialog.getOpenFileName()[0]
            self.data_filename = self.existing_data[self.existing_data.rfind("/") + 1:]
            if self.data_filename == "":
                self.existing_data_button.setText("Input Existing Class I Data (.xlsx)")
            else:
                self.existing_data_button.setText(self.data_filename)
        elif option == 3:
            self.existing_data_2 = QFileDialog.getOpenFileName()[0]
            self.data_filename_2 = self.existing_data_2[self.existing_data_2.rfind("/") + 1:]
            if self.data_filename_2 == "":
                self.existing_data_2_button.setText("Input Existing Class II Data (.xlsx)")
            else:
                self.existing_data_2_button.setText(self.data_filename_2)

    def write_to_textbox(self, text):
        cursor = self.output_textbox.textCursor()
        cursor.movePosition(cursor.End)
        cursor.insertText(text)

    def back_home_auto_epi_collect(self):
        self.stackedWidget.setCurrentWidget(self.options_page)

    def run_auto_epi_collect(self, gene, gene_file, existing_data, existing_data2, mut, i, an, al, ali, g, iso, h, ins,
                             t, ifn,
                             filtering, scoring, pop_cov, mhc_classes):
        if self.worker.run_times > 0:
            self.worker_thread.quit()
            self.worker_thread.wait()
            self.output_textbox.clear()
            self.worker = Worker()
            self.worker_thread = QThread()
            self.worker.moveToThread(self.worker_thread)
            self.worker.update_signal.connect(self.write_to_textbox)
        self.worker_thread.started.connect(
            partial(self.worker.auto_epi_collect, gene, gene_file, existing_data, existing_data2, mut, i, an, al, ali,
                    g, iso, h, ins,
                    t, ifn, filtering, scoring, pop_cov, mhc_classes))
        self.worker_thread.start()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec_())